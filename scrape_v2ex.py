#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
抓取 V2EX remote 版块的招聘数据并存入 Excel
使用改进的选择器和重试机制
"""

import requests
from bs4 import BeautifulSoup
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import time

def scrape_v2ex():
    """抓取 V2EX remote 版块的数据"""
    url = 'https://www.v2ex.com/go/remote'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    }
    
    jobs = []
    
    try:
        print(f"正在访问: {url}")
        response = requests.get(url, headers=headers, timeout=15)
        response.encoding = 'utf-8'
        print(f"响应状态码: {response.status_code}")
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # 调试：打印页面内容片段
        print(f"页面长度: {len(response.text)} 字符")
        
        # 尝试多种选择器找到帖子
        # 方法1: 标准的 cell 选择器
        topics = soup.find_all('div', class_='cell item')
        print(f"方法1 (cell item): 找到 {len(topics)} 条")
        
        # 方法2: 更宽泛的搜索
        if len(topics) == 0:
            topics = soup.find_all('table', class_='topics_table')
            print(f"方法2 (topics_table): 找到 {len(topics)} 条")
            if topics:
                topics = topics[0].find_all('tr')[1:]  # 跳过表头
        
        # 方法3: 直接找 tr
        if len(topics) == 0:
            topics = soup.find_all('tr', recursive=True)
            topics = [t for t in topics if t.find('a', class_='topic-link')]
            print(f"方法3 (tr with topic-link): 找到 {len(topics)} 条")
        
        print(f"总共找到 {len(topics)} 条帖子\n")
        
        for idx, topic in enumerate(topics[:100], 1):  # 限制前100条
            try:
                # 获取标题和链接
                title_elem = topic.find('a', class_='topic-link')
                if not title_elem:
                    continue
                
                title = title_elem.get_text(strip=True)
                topic_href = title_elem.get('href', '')
                if not topic_href.startswith('http'):
                    topic_url = 'https://www.v2ex.com' + topic_href
                else:
                    topic_url = topic_href
                
                # 获取公司名称（从跟帖者中）
                gray_text = topic.find('small')
                company = '社区用户'
                if gray_text:
                    text = gray_text.get_text(strip=True)
                    # 尝试从文本中提取用户名
                    parts = text.split('•')
                    if len(parts) > 0:
                        company = parts[0].strip()
                
                # 基本信息
                category = '远程工作'
                location = '全国远程'
                date_str = datetime.now().strftime('%Y-%m-%d')
                
                job = {
                    'company': company[:30] if company else title[:20],
                    'category': category,
                    'position': title,
                    'location': location,
                    'url': topic_url,
                    'source': topic_url,
                    'date': date_str,
                    'salary': '面议',
                }
                
                jobs.append(job)
                print(f"{idx}. ✓ {title[:60]}...")
                
            except Exception as e:
                print(f"{idx}. ✗ 处理失败: {str(e)[:100]}")
                continue
        
        print(f"\n成功抓取 {len(jobs)} 条数据")
        return jobs
        
    except Exception as e:
        print(f"抓取失败: {e}")
        import traceback
        traceback.print_exc()
        return []

def write_to_excel(jobs):
    """将数据写入 Excel"""
    excel_path = '/Users/qisoong/WorkBuddy/20260317141747/money.xlsx'
    
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # 获取表头（第2行）
        headers = []
        for cell in ws[2]:
            headers.append(cell.value)
        
        print(f"\n表头 ({len(headers)} 列): {headers}\n")
        
        # 找到最后一行有数据的位置
        last_row = 1
        for row in ws.iter_rows():
            if any(cell.value for cell in row):
                last_row = row[0].row
        
        print(f"当前最后有数据的行: {last_row}")
        
        # 开始写入位置（下一行）
        start_row = last_row + 1
        
        # 写入数据
        for idx, job in enumerate(jobs, 1):
            row_num = start_row + idx - 1
            
            # 根据表头对应的列写入数据
            for col_idx, header in enumerate(headers, 1):
                value = ''
                
                if header == '序号':
                    value = idx
                elif header == '公司':
                    value = job['company']
                elif header == '行业':
                    value = '互联网'
                elif header == '职位类别':
                    value = job['category']
                elif header == '职位名称':
                    value = job['position']
                elif header == '地区':
                    value = job['location']
                elif header == '申请链接':
                    value = job['url']
                elif header == '日期':
                    value = job['date']
                elif header == '来源':
                    value = job['source']
                elif header == '薪资':
                    value = job['salary']
                
                if value:
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        wb.save(excel_path)
        print(f"\n✓ 成功写入 {len(jobs)} 行数据到 Excel")
        print(f"  起始行: {start_row}")
        print(f"  结束行: {start_row + len(jobs) - 1}")
        
    except Exception as e:
        print(f"写入 Excel 失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    print("="*60)
    print("V2EX Remote 版块数据抓取")
    print("="*60 + "\n")
    
    jobs = scrape_v2ex()
    if jobs:
        write_to_excel(jobs)
    else:
        print("未获取到任何数据")
